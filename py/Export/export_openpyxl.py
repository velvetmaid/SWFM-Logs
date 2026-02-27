import os
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import math
import re

# --- Load .env ---
load_dotenv()

# --- DB credentials ---
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_NAME = os.getenv("DB_NAME")

# --- Timestamp for file naming ---
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

# --- Connect to DB ---
conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    user=DB_USER,
    password=DB_PASSWORD,
    dbname=DB_NAME
)

# --- Query ---
with open("query.sql", "r") as f:
    query = f.read()

# --- Get data ---
df = pd.read_sql_query(query, conn)

# --- Clean illegal chars ---
def clean_illegal_chars(val):
    if isinstance(val, str):
        return re.sub(r"[\x00-\x1F\x7F]", "", val)
    return val

df = df.applymap(clean_illegal_chars)

# --- Close conn ---
conn.close()

# --- Export Excel (split into multiple sheets if needed) ---
excel_path = f"{current_time}_export.xlsx"
max_rows = 1048576  # Excel limit
usable_rows = max_rows - 1

num_sheets = math.ceil(len(df) / usable_rows)

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    for i in range(num_sheets):
        start_row = i * usable_rows
        end_row = min((i + 1) * usable_rows, len(df))
        df_chunk = df.iloc[start_row:end_row]

        sheet_name = f"Sheet{i+1}"
        df_chunk.to_excel(writer, sheet_name=sheet_name, index=False)


# --- Format Excel ---
wb = load_workbook(excel_path)

for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Bold header + freeze pane
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # Sortable table
    table_range = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"
    table = Table(displayName=f"DataTable_{sheet}", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

# Save
wb.save(excel_path)
print(f"\n Finished -> {excel_path} dengan {num_sheets} sheet")
