import os
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import paramiko
import time
from datetime import datetime

load_dotenv()

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_NAME = os.getenv("DB_NAME")

SFTP_HOST = os.getenv("SFTP_HOST")
SFTP_PORT = int(os.getenv("SFTP_PORT", 22))
SFTP_USER = os.getenv("SFTP_USER")
SFTP_PASS = os.getenv("SFTP_PASSWORD")

LOCAL_ATTACHMENT_FOLDER = "attachments"
os.makedirs(LOCAL_ATTACHMENT_FOLDER, exist_ok=True)

current_time = datetime.now().strftime("%Y%m%d_%H%M%S_")

conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    user=DB_USER,
    password=DB_PASSWORD,
    dbname=DB_NAME
)

query_main = """
select tptg.tx_pm_ticket_genset_id as ticket_no, * from wfm_schema.tx_pm_ticket_genset tptg
inner join wfm_schema.tx_pm_genset_genset tpgg
on tptg.tx_pm_ticket_genset_id = tpgg.tx_pm_ticket_genset_id
left join wfm_schema.tx_site ts 
on tptg.tx_site_id = ts.site_id
where tptg.tx_site_id in ('AGA006',
'BAU026',
'BAU034',
'BLS260',
'BTM093',
'BTX003',
'BTX025',
'COK119',
'DAK003',
'FFK016',
'KBA068',
'KEP003',
'KEP006',
'KKP225',
'KMN010',
'KTP049',
'LBA013',
'LBA015',
'MDO052',
'MRK022',
'MRW006',
'MRW025',
'MSH011',
'MSH013',
'MSH049',
'MSH059',
'MSH070',
'MSH071',
'MSH072',
'MSH096',
'MSH112',
'MSH370',
'MSH383',
'MSH384',
'MSH523',
'MSH624',
'MTW014',
'MTW016',
'NAB024',
'NGP018',
'PGI001',
'PKJ010',
'PKJ125',
'PKJ126',
'PKR515',
'PKR584',
'PKR655',
'PSO022',
'PTS004',
'PTS027',
'RHA060',
'SAK017',
'SAK032',
'SLY009',
'SML014',
'SML150',
'SML254',
'SML258',
'SON129',
'SON417',
'SRU130',
'SSU003',
'STA009',
'STA013',
'STA050',
'TBH006',
'TBH035',
'TBH036',
'TBH042',
'TBH057',
'TBH139',
'TBH647',
'TBH669',
'TBH674',
'TBH680',
'TBH684',
'TBK523',
'TDO039',
'TIM065',
'TIM327',
'TMP002',
'TOB005',
'TOB006',
'TPI005',
'TPI530',
'TPI602',
'TPJ248',
'TPJ249',
'TUL163',
'WSI003',
'WSI075',
'WSI094',
'MSH389',
'MSH425',
'MSH443',
'NLA148',
'NLA152',
'PTS026',
'SML108',
'SML156',
'SML255',
'TPI047',
'WGP015',
'AGA008',
'BEK064',
'BLS119',
'BLS165',
'BLS321',
'BNT025',
'BTI058',
'COH024',
'COH035',
'COI346',
'FFK013',
'KAI120',
'KBA059',
'KBA298',
'KKN224',
'KKP067',
'KKP072',
'KKP217',
'KPA040',
'KPA110',
'KPG325',
'KSN055',
'KSN056',
'KSN150',
'KTP107',
'KTP214',
'LBA090',
'LBJ011',
'MLE066',
'MLI008',
'MLU008',
'MTK004',
'MUL001',
'MUL005',
'NAB023',
'NIK015',
'OKI233',
'OKI408',
'OKI444',
'PBU079',
'PBU083',
'PBU105',
'PBU276',
'PKJ060',
'PKJ109',
'PPN130',
'PTS028',
'RGT214',
'RTA034',
'SAA012',
'SAA083',
'SAA085',
'SAA089',
'SAA091',
'SAA448',
'SAK007',
'SAK014',
'SAK028',
'SBS088',
'SLY030',
'SMA014',
'SMA018',
'SMI128',
'SMI129',
'SMI131',
'SMP272',
'SMP273',
'SMR529',
'SNN016',
'SPT114',
'TBH033',
'TBH075',
'TDO142',
'TPI516',
'TRG177',
'TRG789',
'TRG902',
'BIA019',
'BLS113',
'BTI059',
'JAP646',
'MAB075',
'MME179',
'MRW229',
'MSH085',
'MUL006',
'NLA016',
'NLA211',
'NTK067',
'NTK080',
'PNJ006',
'SAA065',
'SAK023',
'SNN013',
'SNN015',
'TIM355',
'TRG157',
'TUL222',
'TUL328',
'TUL392',
'UPX147',
'SAT111',
'AGR004',
'OKI095',
'SKY138',
'LLG013',
'BIM054',
'BIM190',
'BIM191',
'LBJ093',
'NTG034',
'END015',
'KAI020',
'LBJ013',
'LRT101',
'LWA034',
'MME031',
'MME076',
'MME093',
'MME177',
'NTG005',
'NTG006',
'NTG007',
'NTG027',
'RTG013',
'SBW028',
'WKB008',
'WKB060') and ts.regional_id = 'R09'
order by tptg.tx_pm_ticket_genset_id, tptg.tx_site_id asc
"""
df = pd.read_sql_query(query_main, conn)

selected_columns = [
    'site_id', 'site_name', 'status', 'ticket_no',
    'asset_terdapat_di_site', 'asset_terpasang', 'asset_rusak',
    'asset_aktif', 'asset_dicuri', 'tsel_barcode', 'nama_asset', 'merk_asset',
    'category_asset', 'serial_number', 'asset_owner', 'capacity', 'bonet',
    'merk_mesin', 'merk_generator', 'running_hour_bulan_lalu',
    'running_hour_bulan_ini', 'durasi_pengecekan', 'genset_model',
    'coolent_level', 'coolent_level_keterangan', 'diesel_engine_oil',
    'diesel_engine_oil_keterangan', 'tambah_ganti_oli', 'air_cleaner',
    'air_cleaner_keterangan', 'battery_starter', 'battery_starter_keterangan',
    'fan_belt', 'fan_belt_keterangan', 'sos_bottle', 'sos_bottle_keterangan',
    'fuel_injection_nozzle', 'fuel_injection_nozzle_keterangan', 'inhibitor',
    'inhibitor_keterangan', 'kekencangan_baut', 'kekencangan_baut_keterangan',
    'kabel_accu_connector', 'kabel_accu_connector_keterangan',
    'muffler_genset', 'muffler_genset_keterangan', 'air_accu',
    'air_accu_keterangan', 'accu_voltage', 'accu_voltage_keterangan',
    'mesin', 'mesin_keterangan'
]

location_columns = [col for col in df.columns if "location" in col.lower()]

def get_path_by_guid(guid):
    query = f"""
        SELECT file_path FROM wfm_admin_schema.tx_attachment_reference
        WHERE id = '{guid}'
    """
    df_path = pd.read_sql_query(query, conn)
    return df_path.iloc[0]['file_path'] if not df_path.empty else None

def download_attachment_from_sftp(path_from_db, filename=None, max_retries=3):
    if not filename:
        filename = os.path.basename(path_from_db) + ".jpg"
    local_path = os.path.join(LOCAL_ATTACHMENT_FOLDER, filename)

    if os.path.exists(local_path):
        print(f"✅ File sudah ada, skip download: {filename}")
        return local_path

    for attempt in range(1, max_retries + 1):
        try:
            print(f"🌀 Attempt {attempt}: Connecting to SFTP...")

            transport = paramiko.Transport((SFTP_HOST, SFTP_PORT))
            transport.banner_timeout = 20
            transport.connect(username=SFTP_USER, password=SFTP_PASS)
            transport.set_keepalive(30)

            sftp = paramiko.SFTPClient.from_transport(transport)
            sftp.get(path_from_db, local_path)

            sftp.close()
            transport.close()

            print(f"📥 Downloaded {os.path.basename(path_from_db)} -> {local_path}")
            return local_path

        except Exception as e:
            print(f"❌ Failed to download {path_from_db} (attempt {attempt}): {e}")
            time.sleep(2)

    print(f"💀 Gagal total download {path_from_db} setelah {max_retries} percobaan")
    return None


for col in location_columns:
    for idx, guid in df[col].items():
        if pd.isna(guid):
            continue
        path = get_path_by_guid(guid)
        if path:
            filename = f"{guid}.jpg"
            downloaded = download_attachment_from_sftp(path, filename)
            if downloaded:
                df.loc[idx, f"{col}_filename"] = filename

conn.close()

excel_columns = selected_columns + [
    f"{col}_filename" for col in location_columns if f"{col}_filename" in df.columns
]
df_export = df[excel_columns]

excel_path = current_time + "output_data_with_images.xlsx"
df_export.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active

for cell in ws[1]:
    cell.font = Font(bold=True)
ws.freeze_panes = "A2"

header_map = {}
for cell in ws[1]:
    if isinstance(cell.value, str) and cell.value.endswith("_location_filename"):
        cell.value = cell.value.replace("_location_filename", "")

FIXED_WIDTH = 150
FIXED_HEIGHT = 300

for idx, row in df_export.iterrows():
    row_num = idx + 2
    for col in location_columns:
        filename_col = f"{col}_filename"
        if filename_col in row and pd.notna(row[filename_col]):
            image_path = os.path.join(LOCAL_ATTACHMENT_FOLDER, row[filename_col])
            if os.path.exists(image_path):
                try:
                    resized_path = image_path.replace(".jpg", "_fixed.jpg")
                    with Image.open(image_path) as im:
                        resized_im = im.resize((FIXED_WIDTH, FIXED_HEIGHT), Image.LANCZOS)
                        resized_im.save(resized_path)

                    img = XLImage(resized_path)
                    img.width = FIXED_WIDTH
                    img.height = FIXED_HEIGHT

                    col_index = list(df_export.columns).index(filename_col)
                    col_letter = ws.cell(row=1, column=col_index + 1).column_letter

                    ws[f"{col_letter}{row_num}"].value = None
                    ws.add_image(img, f"{col_letter}{row_num}")

                    ws.column_dimensions[col_letter].width = FIXED_WIDTH / 7.5
                    ws.row_dimensions[row_num].height = FIXED_HEIGHT * 0.75

                except Exception as e:
                    print(f"❌ Gagal masukin gambar {image_path}: {e}")


for col in ws.columns:
    col_letter = col[0].column_letter
    header_val = ws[f"{col_letter}1"].value
    if header_val and not header_val.endswith("_filename"):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col_letter].width = max_len + 2

wb.save(excel_path)
print(f"\n✅ Excel siap! 💪 -> {excel_path}")
