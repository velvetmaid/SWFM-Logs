import os
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

current_time = datetime.now().strftime("%Y%m%d_%H%M%S_")
load_dotenv()

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_NAME = os.getenv("DB_NAME")

conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    user=DB_USER,
    password=DB_PASSWORD,
    dbname=DB_NAME
)

query = """WITH combined_users AS (
    SELECT
        a.employee_name AS name,
        a.email,
        b.area_name,
        c.regional_name,
        d.nop_name,
        e.cluster_name,
        string_agg(DISTINCT t.name::text, ', ') AS role_name,
        'WEB' AS source_user
    FROM
        wfm_schema.tx_user_management a
        LEFT JOIN wfm_schema.tm_area b ON a.area_id = b.area_id
        LEFT JOIN wfm_schema.tm_regional c ON a.regional_id = c.regional_id
        LEFT JOIN wfm_schema.tm_nop d ON a.nop_id = d.nop_id
        LEFT JOIN wfm_schema.tm_cluster e ON a.cluster_id = e.cluster_id
        INNER JOIN wfm_schema.tx_user_role tu ON a.ref_user_id = tu.ref_user_id
        INNER JOIN wfm_schema.tm_user_role t ON tu.role_id = t.tm_user_role_id
    WHERE
        a.is_active = TRUE
        AND a.is_delete = FALSE
    GROUP BY
        a.employee_name, a.email, b.area_name, c.regional_name, d.nop_name, e.cluster_name
    UNION ALL
    SELECT
        a.employee_name AS name,
        a.email,
        b.area_name,
        c.regional_name,
        d.nop_name,
        e.cluster_name,
        string_agg(DISTINCT t.name::text, ', ') AS role_name,
        'MOBILE' AS source_user
    FROM
        wfm_schema.tx_user_mobile_management a
        LEFT JOIN wfm_schema.tm_area b ON a.area_id = b.area_id
        LEFT JOIN wfm_schema.tm_regional c ON a.regional_id = c.regional_id
        LEFT JOIN wfm_schema.tm_nop d ON a.nop_id = d.nop_id
        LEFT JOIN wfm_schema.tm_cluster e ON a.cluster_id = e.cluster_id
        INNER JOIN wfm_schema.mapping_user_mobile_role m ON a.tx_user_mobile_management_id = m.tx_user_mobile_management_id
        INNER JOIN wfm_schema.tm_user_role t ON m.role_id = t.tm_user_role_id
    WHERE
        a.is_active = TRUE
        AND a.is_delete = FALSE
    GROUP BY
        a.employee_name, a.email, b.area_name, c.regional_name, d.nop_name, e.cluster_name
)
SELECT
    MIN(name) AS name,
    email,
    MIN(area_name) AS area_name,
    MIN(regional_name) AS regional_name,
    MIN(nop_name) AS nop_name,
    MIN(cluster_name) AS cluster_name,
    string_agg(DISTINCT role_name, ', ') AS role_name,
    string_agg(DISTINCT source_user, ', ') AS source_user,
    CASE 
        WHEN email ILIKE '%@telkomsel%' THEN TRUE
        ELSE FALSE
    END AS is_telkomsel_user
FROM
    combined_users
GROUP BY
    email
ORDER BY
    name;"""
df = pd.read_sql_query(query, conn)

conn.close()

excel_path =  "SWFM USERS" + current_time + ".xlsx"
df.to_excel(excel_path, index=False)

wb = load_workbook(excel_path)
ws = wb.active

for cell in ws[1]:
    cell.font = Font(bold=True)

ws.freeze_panes = "A2"

ws.auto_filter.ref = ws.dimensions

for col in ws.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_len + 2

wb.save(excel_path)

print("✅ Output Excel 😎 ->", excel_path)
