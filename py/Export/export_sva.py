import os
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

# --- Load .env ---
load_dotenv()

# --- DB credentials ---
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_NAME = os.getenv("DB_NAME")

# --- Timestamp for file naming ---
current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

# --- Connect to DB ---
conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    user=DB_USER,
    password=DB_PASSWORD,
    dbname=DB_NAME
)

# --- Query utama ---
query = """
select tptg.area_name, tptg.regional_name, tptg.nop_name, tsi.scope_item_name as pmg_type,
to_char(schedule_date, 'FMMonth - yyyy') as bulan, tptg.tx_pm_ticket_genset_id as nomor_ticket, tptg.status as status_ticket,
CASE 
  WHEN
    trim(coalesce(tpgs.foto_pergantian_oli_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_pergantian_oli_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_pergantian_filter_oli_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_pergantian_filter_oli_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_pergantian_filter_solar_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_pergantian_filter_solar_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_pergantian_filter_udara_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_pergantian_filter_udara_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_renew_coolant_radiator_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_renew_coolant_radiator_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_renew_all_gasket_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_renew_all_gasket_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_renew_nozzle_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_renew_nozzle_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_renew_fan_belt_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_renew_fan_belt_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_renew_radiator_hose_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_renew_radiator_hose_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_renew_inlet_valve_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_renew_inlet_valve_sesudah_sftp_id, '')) = '' AND
    trim(coalesce(tpgs.foto_renew_exhaust_valve_sebelum_sftp_id, '')) = '' AND trim(coalesce(tpgs.foto_renew_exhaust_valve_sesudah_sftp_id, '')) = ''
  THEN 'Tidak ada foto penggantian spare part'
  WHEN
    trim(coalesce(tpgs.foto_pergantian_oli_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_pergantian_oli_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_pergantian_filter_oli_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_pergantian_filter_oli_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_pergantian_filter_solar_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_pergantian_filter_solar_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_pergantian_filter_udara_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_pergantian_filter_udara_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_renew_coolant_radiator_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_renew_coolant_radiator_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_renew_all_gasket_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_renew_all_gasket_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_renew_nozzle_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_renew_nozzle_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_renew_fan_belt_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_renew_fan_belt_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_renew_radiator_hose_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_renew_radiator_hose_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_renew_inlet_valve_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_renew_inlet_valve_sesudah_sftp_id, '')) != '' OR
    trim(coalesce(tpgs.foto_renew_exhaust_valve_sebelum_sftp_id, '')) != '' OR trim(coalesce(tpgs.foto_renew_exhaust_valve_sesudah_sftp_id, '')) != ''
  THEN 'Ada beberapa foto penggantian spare part'
END AS keterangan,
TRIM(BOTH ', ' FROM
  CASE WHEN trim(tpgs.foto_pergantian_oli_sebelum_sftp_id) != '' OR trim(tpgs.foto_pergantian_oli_sesudah_sftp_id) != '' THEN 'oli, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_pergantian_filter_oli_sebelum_sftp_id) != '' OR trim(tpgs.foto_pergantian_filter_oli_sesudah_sftp_id) != '' THEN 'filter oli, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_pergantian_filter_solar_sebelum_sftp_id) != '' OR trim(tpgs.foto_pergantian_filter_solar_sesudah_sftp_id) != '' THEN 'filter solar, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_pergantian_filter_udara_sebelum_sftp_id) != '' OR trim(tpgs.foto_pergantian_filter_udara_sesudah_sftp_id) != '' THEN 'filter udara, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_renew_coolant_radiator_sebelum_sftp_id) != '' OR trim(tpgs.foto_renew_coolant_radiator_sesudah_sftp_id) != '' THEN 'coolant radiator, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_renew_all_gasket_sebelum_sftp_id) != '' OR trim(tpgs.foto_renew_all_gasket_sesudah_sftp_id) != '' THEN 'gasket, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_renew_nozzle_sebelum_sftp_id) != '' OR trim(tpgs.foto_renew_nozzle_sesudah_sftp_id) != '' THEN 'nozzle, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_renew_fan_belt_sebelum_sftp_id) != '' OR trim(tpgs.foto_renew_fan_belt_sesudah_sftp_id) != '' THEN 'fan belt, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_renew_radiator_hose_sebelum_sftp_id) != '' OR trim(tpgs.foto_renew_radiator_hose_sesudah_sftp_id) != '' THEN 'radiator hose, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_renew_inlet_valve_sebelum_sftp_id) != '' OR trim(tpgs.foto_renew_inlet_valve_sesudah_sftp_id) != '' THEN 'inlet valve, ' ELSE '' END ||
  CASE WHEN trim(tpgs.foto_renew_exhaust_valve_sebelum_sftp_id) != '' OR trim(tpgs.foto_renew_exhaust_valve_sesudah_sftp_id) != '' THEN 'exhaust valve, ' ELSE '' END
) AS spare_part_ada_foto
from wfm_schema.tx_pm_ticket_genset tptg 
left join wfm_schema.tx_pm_genset_sparepart tpgs 
on tptg.tx_pm_ticket_genset_id = tpgs.tx_pm_ticket_genset_id
left join wfm_schema.tm_scope_item tsi 
on tptg.scope_item_id = tsi.scope_item_id 
where (tptg.status = 'CLOSED' OR tptg.status = 'SUBMITTED') and 
tptg.schedule_date >= DATE '2025-01-01' 
AND tptg.schedule_date < DATE '2025-07-01'
"""

# --- Ambil Data ---
df = pd.read_sql_query(query, conn)

# --- Bersihin karakter haram dari Excel ---
import re
def clean_illegal_chars(val):
    if isinstance(val, str):
        return re.sub(r"[\x00-\x1F\x7F]", "", val)
    return val

df = df.applymap(clean_illegal_chars)

# --- Tutup koneksi DB ---
conn.close()

# --- Export ke Excel ---
excel_path = f"{current_time}_export.xlsx"
df.to_excel(excel_path, index=False)

# --- Format Excel ---
wb = load_workbook(excel_path)
ws = wb.active

# Bold header + freeze pane
for cell in ws[1]:
    cell.font = Font(bold=True)
ws.freeze_panes = "A2"

# Tambahin fitur sortable table
table_range = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"
table = Table(displayName="DataTable", ref=table_range)
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

# Simpan
wb.save(excel_path)
print(f"\n✅ EXCEL JADI! 🧾 -> {excel_path}")
