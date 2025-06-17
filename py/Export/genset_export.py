import os
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import paramiko
import time
from datetime import datetime

load_dotenv()

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_NAME = os.getenv("DB_NAME")

SFTP_HOST = os.getenv("SFTP_HOST")
SFTP_PORT = int(os.getenv("SFTP_PORT", 22))
SFTP_USER = os.getenv("SFTP_USER")
SFTP_PASS = os.getenv("SFTP_PASSWORD")

LOCAL_ATTACHMENT_FOLDER = "attachments"
os.makedirs(LOCAL_ATTACHMENT_FOLDER, exist_ok=True)

# Timestamp
current_time = datetime.now().strftime("%Y%m%d_%H%M%S_")

conn = psycopg2.connect(
    host=DB_HOST,
    port=DB_PORT,
    user=DB_USER,
    password=DB_PASSWORD,
    dbname=DB_NAME
)

# Main data
query_main = """
select tptg.tx_pm_ticket_genset_id as ticket_no, * from wfm_schema.tx_pm_ticket_genset tptg
inner join wfm_schema.tx_pm_genset_genset tpgg
on tptg.tx_pm_ticket_genset_id = tpgg.tx_pm_ticket_genset_id
left join wfm_schema.tx_site ts 
on tptg.tx_site_id = ts.site_id
where tptg.tx_site_id in ('AGA006',
'BAU026',
'BAU034',
'WKB060') and ts.regional_id = 'R10'
order by tptg.tx_pm_ticket_genset_id, tptg.tx_site_id asc
"""
df = pd.read_sql_query(query_main, conn)

# --- Ambil semua GUID ---
all_guids = set()
for col in location_columns:
    all_guids.update(df[col].dropna().tolist())

# --- Ambil semua path by GUID  ---
placeholder = ', '.join(f"'{g}'" for g in all_guids)
query = f"""
    SELECT id, file_path FROM wfm_admin_schema.tx_attachment_reference
    WHERE id IN ({placeholder})
"""
guid_df = pd.read_sql_query(query, conn)
guid_to_path = dict(zip(guid_df['id'], guid_df['file_path']))

# --- Download semua file ---
if not os.path.exists(LOCAL_ATTACHMENT_FOLDER):
    os.makedirs(LOCAL_ATTACHMENT_FOLDER)

with paramiko.Transport((SFTP_HOST, SFTP_PORT)) as transport:
    transport.connect(username=SFTP_USER, password=SFTP_PASS)
    sftp = paramiko.SFTPClient.from_transport(transport)

    for col in location_columns:
        for idx, guid in df[col].items():
            if pd.isna(guid): continue
            path = guid_to_path.get(guid)
            filename = f"{guid}.jpg"
            local_path = os.path.join(LOCAL_ATTACHMENT_FOLDER, filename)

            if not os.path.exists(local_path) and path:
                try:
                    sftp.get(path, local_path)
                    print(f"📥 Downloaded {filename}")
                except Exception as e:
                    print(f"❌ Failed to download {path}: {e}")
            else:
                print(f"⚡ Skipped (already exists): {filename}")

            df.loc[idx, f"{col}_filename"] = filename

conn.close()

# --- Simpan awal Excel ---
excel_columns = selected_columns + [
    f"{col}_filename" for col in location_columns if f"{col}_filename" in df.columns
]
df_export = df[excel_columns]

excel_path = current_time + "output_data_with_images.xlsx"
df_export.to_excel(excel_path, index=False)

# --- Masukin gambar ---
FIXED_WIDTH = 150
FIXED_HEIGHT = 300

wb = load_workbook(excel_path)
ws = wb.active

# Header bold + ganti nama
for cell in ws[1]:
    cell.font = Font(bold=True)
    if isinstance(cell.value, str) and cell.value.endswith("_location_filename"):
        cell.value = cell.value.replace("_location_filename", "")

ws.freeze_panes = "A2"

# --- Tambahin gambar ---
for idx, row in df_export.iterrows():
    row_num = idx + 2
    for col in location_columns:
        filename_col = f"{col}_filename"
        if filename_col in row and pd.notna(row[filename_col]):
            image_path = os.path.join(LOCAL_ATTACHMENT_FOLDER, row[filename_col])
            if os.path.exists(image_path):
                try:
                    resized_path = image_path.replace(".jpg", "_fixed.jpg")
                    if not os.path.exists(resized_path):
                        with Image.open(image_path) as im:
                            resized_im = im.resize((FIXED_WIDTH, FIXED_HEIGHT), Image.LANCZOS)
                            resized_im.save(resized_path)

                    img = XLImage(resized_path)
                    img.width = FIXED_WIDTH
                    img.height = FIXED_HEIGHT

                    col_index = list(df_export.columns).index(filename_col)
                    col_letter = ws.cell(row=1, column=col_index + 1).column_letter

                    ws[f"{col_letter}{row_num}"].value = None
                    ws.add_image(img, f"{col_letter}{row_num}")
                    ws.column_dimensions[col_letter].width = FIXED_WIDTH / 7.5
                    ws.row_dimensions[row_num].height = FIXED_HEIGHT * 0.75

                except Exception as e:
                    print(f"❌ Gagal resize/gambar {image_path}: {e}")

# --- Auto width non-image column ---
for col in ws.columns:
    col_letter = col[0].column_letter
    header_val = ws[f"{col_letter}1"].value
    if header_val and not str(header_val).endswith("_filename"):
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col_letter].width = max_len + 2

# --- Simpan Excel final ---
wb.save(excel_path)
print(f"\n✅ Finished -> {excel_path}")